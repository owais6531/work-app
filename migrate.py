"""One-time (re-runnable) migration: Client-Master-List.xlsx -> office.db (SQLite).

Re-running this wipes and rebuilds clients/client_links/tasks tables from the
xlsx, so the xlsx stays the "import source" until the dashboard fully takes
over as primary. Safe to re-run any time to re-sync from a freshly edited xlsx.
"""
import sqlite3
import datetime
import openpyxl

XLSX_PATH = r"C:\Users\owais\My Drive\Owais\Online work\claude employee office tasks\Client-Master-List.xlsx"
DB_PATH = r"C:\Users\owais\OfficeDashboard\office.db"

CATEGORY_COLS = list(range(4, 17))  # D..P: Income Tax 2020..2026, Sales Tax, WHT, Monthly Task, Registration/SECP/Temp, Other

SCHEMA = """
CREATE TABLE IF NOT EXISTS clients (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    ntn TEXT,
    group_family TEXT,
    contact_info TEXT,
    registration_status TEXT,
    last_enriched TEXT,
    enrichment_notes TEXT,
    status_notes TEXT,
    master_list_row INTEGER UNIQUE
);
CREATE TABLE IF NOT EXISTS client_links (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    client_id INTEGER NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
    category TEXT,
    link_text TEXT,
    link_target TEXT
);
CREATE TABLE IF NOT EXISTS tasks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    client_id INTEGER REFERENCES clients(id) ON DELETE SET NULL,
    client_name_raw TEXT,
    task_type TEXT,
    period TEXT,
    status TEXT DEFAULT 'Pending',
    priority TEXT,
    due_date TEXT,
    blocked_on TEXT,
    plan_day INTEGER,
    owner TEXT,
    notes TEXT,
    created_at TEXT,
    updated_at TEXT,
    completed_at TEXT
);
CREATE INDEX IF NOT EXISTS idx_tasks_status ON tasks(status);
CREATE INDEX IF NOT EXISTS idx_tasks_priority ON tasks(priority);
CREATE INDEX IF NOT EXISTS idx_tasks_due ON tasks(due_date);
CREATE INDEX IF NOT EXISTS idx_clients_name ON clients(name);
"""

def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    master = wb["Client Master List"]
    headers = [c.value for c in master[1]]

    con = sqlite3.connect(DB_PATH)
    con.executescript(SCHEMA)
    cur = con.cursor()

    cur.execute("DELETE FROM tasks")
    cur.execute("DELETE FROM client_links")
    cur.execute("DELETE FROM clients")

    now = datetime.datetime.now().isoformat(timespec="seconds")
    row_to_client_id = {}

    for r in range(2, master.max_row + 1):
        name = master.cell(row=r, column=1).value
        if not name:
            continue
        ntn = master.cell(row=r, column=2).value
        group = master.cell(row=r, column=3).value
        status_notes = master.cell(row=r, column=16).value
        contact_info = master.cell(row=r, column=17).value if master.max_column >= 17 else None
        reg_status = master.cell(row=r, column=18).value if master.max_column >= 18 else None
        last_enriched = master.cell(row=r, column=19).value if master.max_column >= 19 else None
        enrich_notes = master.cell(row=r, column=20).value if master.max_column >= 20 else None

        cur.execute(
            "INSERT INTO clients (name, ntn, group_family, contact_info, registration_status, "
            "last_enriched, enrichment_notes, status_notes, master_list_row) VALUES (?,?,?,?,?,?,?,?,?)",
            (str(name).strip(), ntn, group, contact_info, reg_status, last_enriched, enrich_notes, status_notes, r),
        )
        client_id = cur.lastrowid
        row_to_client_id[r] = client_id

        for col in CATEGORY_COLS:
            cell = master.cell(row=r, column=col)
            if cell.value is None and cell.hyperlink is None:
                continue
            link_target = cell.hyperlink.target if cell.hyperlink else None
            cur.execute(
                "INSERT INTO client_links (client_id, category, link_text, link_target) VALUES (?,?,?,?)",
                (client_id, headers[col - 1], str(cell.value) if cell.value is not None else None, link_target),
            )

    task_count = 0
    if "Work Tracker" in wb.sheetnames:
        wt = wb["Work Tracker"]
        for r in range(2, wt.max_row + 1):
            client_name = wt.cell(row=r, column=1).value
            if not client_name:
                continue
            ntn = wt.cell(row=r, column=2).value
            task_type = wt.cell(row=r, column=3).value
            period = wt.cell(row=r, column=4).value
            status = wt.cell(row=r, column=5).value or "Pending"
            priority = wt.cell(row=r, column=6).value
            due_date = wt.cell(row=r, column=7).value
            blocked_on = wt.cell(row=r, column=8).value
            plan_day = wt.cell(row=r, column=9).value
            notes = wt.cell(row=r, column=10).value
            match_text = wt.cell(row=r, column=11).value

            client_id = None
            if match_text and str(match_text).startswith("Row "):
                try:
                    matched_row = int(str(match_text).split("Row ")[1])
                    client_id = row_to_client_id.get(matched_row)
                except (ValueError, IndexError):
                    client_id = None

            owner = "Umair" if blocked_on else "Claude"

            cur.execute(
                "INSERT INTO tasks (client_id, client_name_raw, task_type, period, status, priority, "
                "due_date, blocked_on, plan_day, owner, notes, created_at, updated_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (client_id, str(client_name), task_type, period, status, priority,
                 str(due_date) if due_date else None, blocked_on, plan_day, owner, notes, now, now),
            )
            task_count += 1

    con.commit()
    con.close()
    print(f"Migrated {len(row_to_client_id)} clients, {task_count} tasks into {DB_PATH}")

if __name__ == "__main__":
    main()
