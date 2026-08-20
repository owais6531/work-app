"""Regenerate Client-Master-List.xlsx (both sheets) from office.db so staff
who prefer Excel/Drive always see a current mirror of the dashboard data.
Overwrites the sheets in place; run after dashboard edits, or via the
dashboard's "Export to Excel" button.
"""
import os
import libsql
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from turso_config import TURSO_URL, TURSO_AUTH_TOKEN

HERE = os.path.dirname(os.path.abspath(__file__))
DRIVE_ROOT = os.path.join(os.path.expanduser("~"), "My Drive", "Owais", "Online work", "claude employee office tasks")

REPLICA_PATH = os.path.join(HERE, "office_replica.db")
XLSX_PATH = os.path.join(DRIVE_ROOT, "Client-Master-List.xlsx")


def rows_as_dicts(cur):
    cols = [d[0] for d in cur.description] if cur.description else []
    return [dict(zip(cols, row)) for row in cur.fetchall()]

MASTER_HEADERS = [
    "Client Name", "NTN / STRN / ID", "Group / Family",
    "Income Tax 2020", "Income Tax 2021", "Income Tax 2022", "Income Tax 2023",
    "Income Tax 2024", "Income Tax 2025", "Income Tax 2026",
    "Sales Tax", "WHT Statements", "Monthly Task", "Registration / SECP / Temp",
    "Other / Employee Copies", "Status / Notes",
    "Contact Info (Email/Phone)", "Registration Status (ATL/SRB/etc)",
    "Last Enriched", "Enrichment Notes",
]
TRACKER_HEADERS = [
    "Client Name", "NTN / STRN", "Task Type", "Period", "Status", "Priority",
    "Due Date", "Blocked On", "Plan Day (1-5)", "Notes / Next Action", "Master List Row",
]
PRIORITY_FILL = {
    "URGENT-OVERDUE": "C00000", "URGENT": "E26B0A", "URGENT-VERIFY DATE": "E26B0A", "BLOCKED": "808080",
}


def header_row(ws, headers):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"


def main():
    con = libsql.connect(REPLICA_PATH, sync_url=TURSO_URL, auth_token=TURSO_AUTH_TOKEN)
    con.sync()

    wb = openpyxl.load_workbook(XLSX_PATH)

    for name in ("Client Master List", "Work Tracker"):
        if name in wb.sheetnames:
            del wb[name]
    ws_master = wb.create_sheet("Client Master List", 0)
    ws_tracker = wb.create_sheet("Work Tracker", 1)
    header_row(ws_master, MASTER_HEADERS)
    header_row(ws_tracker, TRACKER_HEADERS)

    clients = rows_as_dicts(con.execute("SELECT * FROM clients ORDER BY id"))
    row_by_client_id = {}
    r = 2
    for c in clients:
        ws_master.cell(row=r, column=1, value=c["name"])
        ws_master.cell(row=r, column=2, value=c["ntn"])
        ws_master.cell(row=r, column=3, value=c["group_family"])
        links = rows_as_dicts(con.execute(
            "SELECT category, link_text, link_target FROM client_links WHERE client_id = ?", (c["id"],)
        ))
        col_by_header = {h: i + 1 for i, h in enumerate(MASTER_HEADERS)}
        for link in links:
            col = col_by_header.get(link["category"])
            if not col:
                continue
            cell = ws_master.cell(row=r, column=col, value=link["link_text"])
            if link["link_target"]:
                cell.hyperlink = link["link_target"]
                cell.font = Font(color="0563C1", underline="single")
        ws_master.cell(row=r, column=16, value=c["status_notes"])
        ws_master.cell(row=r, column=17, value=c["contact_info"])
        ws_master.cell(row=r, column=18, value=c["registration_status"])
        ws_master.cell(row=r, column=19, value=c["last_enriched"])
        ws_master.cell(row=r, column=20, value=c["enrichment_notes"])
        row_by_client_id[c["id"]] = r
        r += 1

    tasks = rows_as_dicts(con.execute("SELECT * FROM tasks ORDER BY id"))
    r = 2
    for t in tasks:
        client_name = t["client_name_raw"]
        master_row = row_by_client_id.get(t["client_id"]) if t["client_id"] else None
        ws_tracker.cell(row=r, column=1, value=client_name)
        ws_tracker.cell(row=r, column=2, value=None)
        ws_tracker.cell(row=r, column=3, value=t["task_type"])
        ws_tracker.cell(row=r, column=4, value=t["period"])
        ws_tracker.cell(row=r, column=5, value=t["status"])
        pcell = ws_tracker.cell(row=r, column=6, value=t["priority"])
        color = PRIORITY_FILL.get(t["priority"])
        if color:
            pcell.fill = PatternFill("solid", fgColor=color)
            pcell.font = Font(color="FFFFFF", bold=True)
        ws_tracker.cell(row=r, column=7, value=t["due_date"])
        ws_tracker.cell(row=r, column=8, value=t["blocked_on"])
        ws_tracker.cell(row=r, column=9, value=t["plan_day"])
        ws_tracker.cell(row=r, column=10, value=t["notes"])
        if master_row:
            link_cell = ws_tracker.cell(row=r, column=11, value=f"Row {master_row}")
            link_cell.hyperlink = f"#'Client Master List'!A{master_row}"
            link_cell.font = Font(color="0563C1", underline="single")
        else:
            ws_tracker.cell(row=r, column=11, value="NOT MATCHED - verify client name")
        r += 1

    for ws, widths in (
        (ws_master, [28, 16, 14] + [18] * 12 + [40, 26, 26, 14, 30]),
        (ws_tracker, [28, 16, 26, 20, 12, 18, 14, 22, 8, 60, 22]),
    ):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=(cell.column == 10 and ws.title == "Work Tracker"))

    wb.save(XLSX_PATH)
    con.close()
    print(f"Exported {len(clients)} clients, {len(tasks)} tasks to {XLSX_PATH}")


if __name__ == "__main__":
    main()
